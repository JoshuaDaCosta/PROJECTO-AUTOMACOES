import tkinter as tk
from tkinter import ttk, messagebox
import openpyxl as pp
import os

# ---------- Configurações ----------
janela = tk.Tk()
janela.title("Vendas de Produtos")
janela.geometry("500x400")

file = "vendas_de_produtos.xlsx"
file_novo = "dados_novos.xlsx"

# ---------- Função para ler categorias do Excel ----------
def ler(file):
    if not os.path.exists(file):
        return []
    livro = pp.load_workbook(file)
    vendas = livro['vendas']
    lista_categoria = set()  # set pra pegar valores únicos
    for linha in vendas.iter_cols(min_col=4, values_only=True):
        for item in linha:
            if item:  # evita None
                lista_categoria.add(item)
    return list(lista_categoria)

# ---------- Função para enviar dados ----------
def enviar():
    nome = cliente_entrada.get()
    produto = produto_entrada.get()
    quantidade = quantidade_entrada.get()
    categoria_selecionada = categoria.get()

    if nome and produto and quantidade and categoria_selecionada:
        try:
            # Cria workbook novo se não existir
            if os.path.exists(file_novo):
                livro = pp.load_workbook(file_novo)
                vendas = livro.active
            else:
                livro = pp.Workbook()
                vendas = livro.active
                vendas.title = "Vendas"
                vendas.append(["Cliente", "Produto", "Quantidade", "Categoria"])

            # Adiciona os dados
            vendas.append([nome, produto, quantidade, categoria_selecionada])
            livro.save(file_novo)
            messagebox.showinfo("Sucesso", "Dados salvos com sucesso!")
            # Limpar entradas
            cliente_entrada.delete(0, tk.END)
            produto_entrada.delete(0, tk.END)
            quantidade_entrada.delete(0, tk.END)
            categoria.delete(0, tk.END)

        except Exception as e:
            messagebox.showerror("Erro", f"Deu um erro: {e}")
    else:
        messagebox.showwarning("Erro", "Preencha todos os campos!")

# ---------- Frames ----------
titulo_frame = tk.Frame(janela, bg="lightgrey", padx=10, pady=10)
titulo_frame.pack(fill="x")

tk.Label(titulo_frame, text="CADASTRAMENTO", font=("Arial", 16)).pack(pady=10)

conteiner_frame = tk.Frame(janela, bg="lightgreen", padx=10, pady=10)
conteiner_frame.pack(fill="both", expand=True)

# ---------- Labels e Entries ----------
tk.Label(conteiner_frame, text="Cliente").grid(row=0, column=0, padx=10, pady=5, sticky="w")
cliente_entrada = tk.Entry(conteiner_frame)
cliente_entrada.grid(row=1, column=0, padx=10, pady=5)

tk.Label(conteiner_frame, text="Produto").grid(row=2, column=0, padx=10, pady=5, sticky="w")
produto_entrada = tk.Entry(conteiner_frame)
produto_entrada.grid(row=3, column=0, padx=10, pady=5)

tk.Label(conteiner_frame, text="Quantidade").grid(row=4, column=0, padx=10, pady=5, sticky="w")
quantidade_entrada = tk.Entry(conteiner_frame)
quantidade_entrada.grid(row=5, column=0, padx=10, pady=5)
 
# ---------- Combobox ----------
tk.Label(conteiner_frame, text="Escolhe uma categoria:").grid(row=0, column=1, padx=10, pady=5, sticky="w")
opcoes = ler(file)
categoria = ttk.Combobox(conteiner_frame, values=opcoes)
categoria.grid(row=1, column=1, padx=10, pady=5)
if opcoes:
    categoria.current(0)==""

# ---------- Botão ----------
botao = tk.Button(conteiner_frame, text="Salvar os dados", command=enviar)
botao.grid(row=5, column=1, padx=10, pady=10)

# ---------- Inicia app ----------
janela.mainloop()
